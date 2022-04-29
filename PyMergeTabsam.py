# -*- coding: utf-8 -*-
# PyMergeTabsam
# @author: sszgrm

import json
import pandas as pd
import datetime
import shutil
import openpyxl
from openpyxl.styles import Font
from copy import copy


# Leere Listen vorbereiten
data_files  = pd.DataFrame([],dtype=pd.StringDtype())
data_sheets = pd.DataFrame([],dtype=pd.StringDtype())

# Global variable from configuration
path_input = ""
filename_output = ""
row_start = 9
row_end = 0

# Function tolog
# Write logging information
def tolog(level, text):
  dateTimeObj = datetime.datetime.now()
  timestamp = dateTimeObj.strftime("%Y-%m-%d %H:%M:%S%z")
  print(level + " (" + timestamp + "): " + text)


# Function read_config
# Read the configuration file
def read_config():
  with open('config.json', 'r', encoding="utf-8") as f:
    config = json.load(f)
    list_files = []
    list_sheets = []
    global data_files, data_sheets, path_input, filename_output
    
    path_input  = config['path_input']
    filename_output = config['filename_output']
    
    for key in config:
      conf_value = config[key]
      if key == "files":
        # The files are provided as a list. Read them an add them do the Dataframe data_files
        for i in range(len(conf_value)):
          files_elem = conf_value[i]
          pk = i+1
          input_fullpath = path_input + "/" + files_elem["input_filename"]
          elem_list_files = [pk, files_elem["title"], input_fullpath, files_elem["position"]]
          list_files.append(elem_list_files)
          data_files = pd.DataFrame(list_files, columns = ['id', 'title' , 'input_path', 'position'])
      if key == "sheets":
        # The sheets to process are provided as a list. Read them an add them do the Dataframe data_sheets
        for i in range(len(conf_value)):
          sheets_elem = conf_value[i]
          pk = i+1
          elem_list_sheets = [pk, sheets_elem["code"], sheets_elem["title"], sheets_elem["column"]]
          list_sheets.append(elem_list_sheets)
          data_sheets = pd.DataFrame(list_sheets, columns = ['id', 'code', 'title', 'column'])


# Function create_tabsam 
# Create and generate the destination excel files
# If the destination files already exists, they will be overwritten
def create_tabsam():
  global row_start, row_end, filename_output

  # Create empty output file based on template
  tolog("INFO", "Writing output to: " + filename_output)
  shutil.copy('VorlageTabsam.xlsx', filename_output)

  sheet_id = 0
  for sheet_i, sheet_row in data_sheets.iterrows():
    sheet_id = sheet_row['id']
    rc = 0
    for file_i, file_row in data_files.iterrows():
      file_id = file_row['id']

      if file_row['position']=="1":
        rc = prepare_table(file_row, sheet_row)
      else:
        if rc == 0:
          rc = merge_table(file_row, sheet_row)
          # Ignore merge errors
          rc = 0
        else:
          # No further processing due to error in primary column
          pass
    row_start = row_end + 3


# Create a new table with header and index column
def prepare_table(file_row, sheet_row):
  global row_start, row_end, filename_output
  return_code = 0

  # Opening the destination xlsx and create the new worksheet
  dest_wb = openpyxl.load_workbook(filename_output)
  dest_ws = dest_wb["T_1"]
  
  # Write table title and set font 
  table_title = sheet_row['code'] + " " + sheet_row['title']
  dest_ws.cell(row=row_start, column=1).value = table_title
  dest_ws.cell(row=row_start, column=1).font = Font(name='Arial', size=8)
  row_start += 1

  # Opening the source xlsx
  source_xlsx = file_row['input_path']
  source_wb = openpyxl.load_workbook(source_xlsx)
  worksheet = sheet_row['code']
  
  # Check if worksheet exists
  if worksheet not in source_wb.sheetnames:
    tolog("ERROR", "Worksheet " + worksheet + " does not exist in " + source_xlsx)
    return_code = 2
  else:
    source_ws = source_wb[worksheet]
    
    row_source = 10
    # Copy the top left cell of the primary table
    topleft_cell = source_ws.cell(row=row_source, column=1)
    dest_ws.cell(row=row_start, column=1).value = topleft_cell.value
    dest_ws.cell(row=row_start, column=1).font  = copy(topleft_cell.font)
    
    # Find the single relevant column
    column_pos = 2
    relevant_name = sheet_row['column']
    # Special case "Ganze Stadt"
    if relevant_name=="Wohnsiedlung" and file_row['title']=="Ganze Stadt":
      relevant_name = "Stadt Zürich"
    while True:
      scan_cell = source_ws.cell(row=row_source, column=column_pos)
      # Convert numeric header cells (eg. years) to character
      if isinstance(scan_cell.value, int):
        scan_cell_value = str(scan_cell.value)
      else:
        scan_cell_value = scan_cell.value
      if scan_cell_value == relevant_name:
        break
      column_pos += 1
      if column_pos > 20:
        column_pos = 0
        tolog("ERROR", "Column " + relevant_name + " does not exist in worksheet " + worksheet + " of " + source_xlsx)
        return_code = 2
        break
    
    # Copy the the header column and the single relevant column
    if column_pos > 0:
      row_target = row_start;
      # Set a title as the column header for the single relevant column
      dest_ws.cell(row=row_start, column=2).value = file_row['title']
      # Apply the font settings of the top left column
      dest_ws.cell(row=row_start, column=2).font  = copy(topleft_cell.font)
      
      while True:
        row_source += 1
        row_target += 1
        head_cell = source_ws.cell(row=row_source, column=1)
        if head_cell.value is None:
          # End of data rows
          break
        if row_source > 100:
          tolog("ERROR", "No end of header column found in worksheet " + worksheet + " of " + source_xlsx)
          return_code = 2
          break
        # copy header cell value, font and alignment
        dest_ws.cell(row=row_target, column=1).value = head_cell.value
        dest_ws.cell(row=row_target, column=1).font  = copy(head_cell.font)
        dest_ws.cell(row=row_target, column=1).alignment = copy(head_cell.alignment)
        # copy data cell value, font, alignment and number format
        data_cell = source_ws.cell(row=row_source, column=column_pos)
        # Special case "Ganze Stadt"
        if relevant_name=="2021" and file_row['title']=="Ganze Stadt":
          pass
        else:
          dest_ws.cell(row=row_target, column=2).value = data_cell.value
          dest_ws.cell(row=row_target, column=2).font  = copy(data_cell.font)
          dest_ws.cell(row=row_target, column=2).alignment = copy(data_cell.alignment)
          dest_ws.cell(row=row_target, column=2).number_format = copy(data_cell.number_format)
        row_end = row_target
    
  dest_wb.save(filename_output)
  return return_code


# merge table to existing one
def merge_table(file_row, sheet_row):
  global row_start, row_end, filename_output
  return_code = 0

  # Opening the destination xlsx 
  dest_wb = openpyxl.load_workbook(filename_output)
  dest_ws = dest_wb["T_1"]
  
  # Opening the source xlsx
  source_xlsx = file_row['input_path']
  source_wb = openpyxl.load_workbook(source_xlsx)
  worksheet = sheet_row['code']
  
  # Check if worksheet exists
  if worksheet not in source_wb.sheetnames:
    tolog("ERROR", "Worksheet " + worksheet + " does not exist in " + source_xlsx)
    return_code = 2
  else:
    source_ws = source_wb[worksheet]
    
    row_source = 10
    topleft_cell = source_ws.cell(row=row_source, column=1)
    row_last = row_source + row_end - row_start

    # Find the single relevant column
    column_pos = 2
    while True:
      scan_cell = source_ws.cell(row=row_source, column=column_pos)
      # Convert numeric header cells (eg. years) to character
      if isinstance(scan_cell.value, int):
        scan_cell_value = str(scan_cell.value)
      else:
        scan_cell_value = scan_cell.value
      if scan_cell_value == sheet_row['column']:
        break
      column_pos += 1
      if column_pos > 20:
        column_pos = 0
        tolog("ERROR", "Column " + sheet_row['column'] + " does not exist in worksheet " + worksheet + " of " + source_xlsx)
        return_code = 2
        break
    
    # Compare the header column and copy the data of the single relevant column
    if column_pos > 0:
      target_col_pos = 1 + int(file_row['position'])
      # Set a title as the column header for the single relevant column
      dest_ws.cell(row=row_start, column=target_col_pos).value = file_row['title']
      # Apply the font settings of the top left column
      dest_ws.cell(row=row_start, column=target_col_pos).font  = copy(topleft_cell.font)
      
      while True:
        row_source += 1
        row_target = row_start
        head_cell = source_ws.cell(row=row_source, column=1)
        if head_cell.value is None:
          # End of data rows
          break
        
        while True:
          row_target += 1
          reference_cell_value = dest_ws.cell(row=row_target, column=1).value
          if reference_cell_value is None:
            # Merging table has more rows than primary one. Record needs to be added.
            # copy header cell value, font and alignment
            dest_ws.cell(row=row_target, column=1).value = head_cell.value
            dest_ws.cell(row=row_target, column=1).font  = copy(head_cell.font)
            dest_ws.cell(row=row_target, column=1).alignment = copy(head_cell.alignment)
            tolog("WARNING", "New row was added add the end of the table " + sheet_row['code'] + " for value " + head_cell.value + " from " + worksheet + " of " + source_xlsx)
            row_end += 1
            # Set reference value to activate the data copy
            reference_cell_value = dest_ws.cell(row=row_target, column=1).value
          if reference_cell_value == head_cell.value:
            # copy data cell value, font, alignment and number format
            data_cell = source_ws.cell(row=row_source, column=column_pos)
            dest_ws.cell(row=row_target, column=target_col_pos).value = data_cell.value
            dest_ws.cell(row=row_target, column=target_col_pos).font  = copy(data_cell.font)
            dest_ws.cell(row=row_target, column=target_col_pos).alignment = copy(data_cell.alignment)
            dest_ws.cell(row=row_target, column=target_col_pos).number_format = copy(data_cell.number_format)
            value_found=1
            break
          
  dest_wb.save(filename_output)
  return return_code


# Main progam
def main():
  tolog("INFO", "Read the configuration")
  read_config()
  
  tolog("INFO", "Loop over the input sheets and files and merge the tables")
  create_tabsam()


# Execute main of PyTabsam
if __name__ == '__main__':
  main()
